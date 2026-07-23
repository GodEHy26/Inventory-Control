#!/usr/bin/env python3
"""
Extract seed data for the Inventory Control app from the
"INVENTORY STOCK LIST" Excel workbook.

Usage:
    python3 scripts/extract_from_xlsx.py <path-to-xlsx> <output-json>

Produces a single seed.json containing:
    warehouses, items, stock (per warehouse), incoming shipments
    (in-transit + ordered), auto-generated BOMs for BNW/BNWW sets,
    and an exceptions report for anything that needed judgement.

Column mapping (per the July 2026 workbook):
    - main inventory column   -> on_hand at WH-MAIN
    - "INVENTORY SAIF ZONE"   -> on_hand at WH-SAIF
    - INTRANSIT columns       -> incoming shipments, status in_transit
    - ORDERED columns         -> incoming shipments, status ordered
    - SALES / PO columns      -> ignored for stock (already reflected in
                                 physical counts); listed in exceptions
    - CLOSING                 -> reference only (projection, not on-hand)
"""
import json, re, sys, datetime
from openpyxl import load_workbook

SHEETS = {
    "A325M FT BNW HDG C8":  dict(item_type="assembly", category="Bolt-Nut-Washer Set", grade="A325M", coating="HDG", name="A325M T-1 FT Bolt-Nut-Washer Set HDG {size}", washers=1),
    "A325M FT BNWW HDG":    dict(item_type="assembly", category="Bolt-Nut-Washer Set", grade="A325M", coating="HDG", name="A325M T-1 FT Bolt-Nut-2xWasher Set HDG {size}", washers=2),
    "A325M HT HDG":         dict(item_type="part", category="Bolt", grade="A325M", coating="HDG", name="A325M T-1 Heavy Hex Bolt (half thread) HDG {size}"),
    "A325M FT HDG":         dict(item_type="part", category="Bolt", grade="A325M", coating="HDG", name="A325M T-1 Heavy Hex Full Thread Bolt HDG {size}"),
    "A563M HDG":            dict(item_type="part", category="Nut", grade="A563M 10S", coating="HDG", name="A563M 10S Heavy Hex Nut HDG {size}"),
    "F436M HDG":            dict(item_type="part", category="Washer", grade="F436M", coating="HDG", name="F436M Hardened Washer HDG {size}"),
    "A490M FT BLACK":       dict(item_type="part", category="Bolt", grade="A490M", coating="Black", name="A490M T-1 Heavy Hex Full Thread Bolt Black {size}"),
    "A563M BLACK":          dict(item_type="part", category="Nut", grade="A563M 10S", coating="Black", name="A563M 10S Heavy Hex Nut Black {size}"),
    "F436M BLACK":          dict(item_type="part", category="Washer", grade="F436M", coating="Black", name="F436M Hardened Washer Black {size}"),
    "A563 HDG":             dict(item_type="part", category="Nut", grade="A563", coating="HDG", name="A563 Heavy Hex Nut HDG {size}", prefix="NA563GZI"),
    "A325 FT HDG":          dict(item_type="part", category="Bolt", grade="A325", coating="HDG", name="A325 Heavy Hex FT Bolt HDG {size}", prefix="BA325GI"),
    "F436 HDG":             dict(item_type="part", category="Washer", grade="F436", coating="HDG", name="F436 Hardened Washer HDG {size}", prefix="WF436GI"),
    "A490 FT BLACK":        dict(item_type="part", category="Bolt", grade="A490", coating="Black", name="A490 Heavy Hex FT Bolt Black {size}", prefix="BA490BI"),
    "A563 BLACK":           dict(item_type="part", category="Nut", grade="A563", coating="Black", name="A563 Heavy Hex Nut Black {size}", prefix="NA563BI"),
    "F436 BLACK":           dict(item_type="part", category="Washer", grade="F436", coating="Black", name="F436 Hardened Washer Black {size}", prefix="WF436BI"),
    "M12 YZP":              dict(item_type="part", category="DIN Fastener", grade="", coating="YZP", name="{code} {size}", code_is_name=True, prefix="DINYZP"),
    "ZIN PALAT WHASHER":    dict(item_type="part", category="Washer", grade="", coating="Zinc Plated", name="Zinc Plated Washer {size}", prefix="WZP"),
    "DIN 125":              dict(item_type="part", category="Washer", grade="DIN125A", coating="HDG", name="{code} {size}", code_is_name=True, prefix="WDIN125"),
    "DIN 934 CL 8":         dict(item_type="part", category="Nut", grade="DIN934 CL8", coating="HDG", name="DIN934 CL8 Nut HDG {size}"),
    "ANCHOR BOLTS":         dict(item_type="part", category="Anchor Bolt", grade="", coating="", name="{code} {size}", code_is_name=True, prefix="ANC"),
    "SAGRODS":              dict(item_type="part", category="Sag Rod", grade="", coating="", name="Sag Rod {size}", prefix="SAG"),
    "ROUND BAR":            dict(item_type="raw", category="Round Bar / Rod", coating="", name="Round Bar {grade} {size}", prefix="RB", uom="each"),
    "SCREW":                dict(item_type="part", category="Self Drill Screw", grade="", coating="", name="Self Drill Screw {size}", prefix="SDS"),
}

def clean(s):
    return re.sub(r"\s+", " ", str(s)).strip() if s is not None else ""

def slug(s):
    return re.sub(r"[^A-Z0-9]+", "", clean(s).upper())[:24]

def is_num(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)

def col_kind(h):
    H = clean(h).upper()
    if not H: return None
    if "SAIF" in H: return "saif"
    if H.startswith("CLOSING"): return "closing"
    if "INTRANSIT" in H or "IN TRANSIT" in H or "NEWLY" in H: return "intransit"
    if H.startswith("ORDER") or "TO BE" in H or H == "BNW": return "ordered"
    if H.startswith("SALES") or H == "PO" or "SHIPPED TO" in H: return "sales"
    if "INVENTORY" in H or "BEGINNING" in H or H.startswith("TOTAL AS") or H == "QTY": return "main"
    if H == "SIZE": return "size"
    if H == "CODE": return "code"
    if H == "GRADE": return "grade"
    return None

def find_header(rows):
    for i, r in enumerate(rows[:12]):
        vals = [clean(c).upper() for c in r]
        if "CODE" in vals and "SIZE" in vals: return i
        if "SIZE" in vals and ("QTY" in vals or "GRADE" in vals or any("INVENTORY" in v for v in vals)): return i
    return None

def meta_lookup(rows, header_i, col):
    """Grab ETA / party / invoice ref for a shipment column from the rows above the header."""
    meta = {}
    for r in rows[:header_i]:
        label = ""
        for c in r[:8]:
            cl = clean(c).upper().rstrip(":")
            if cl in ("ETA","DATE","NAME","FACTORY","PO","BL","B/L","INV NO"): label = cl; break
        if not label: continue
        v = r[col] if col < len(r) else None
        if v is None: continue
        if isinstance(v, datetime.datetime): v = v.date().isoformat()
        v = clean(v)
        if not v: continue
        if label in ("ETA","DATE") and re.match(r"\d{4}-\d{2}-\d{2}", v): meta["eta"] = v[:10]
        elif label in ("NAME","FACTORY"): meta["party"] = v
        elif label in ("PO","INV NO"): meta.setdefault("ref", v)
        elif label in ("BL","B/L"): meta.setdefault("bl", v)
    return meta

def main(xlsx_path, out_path):
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    items, stock, shipments, boms, exceptions = {}, [], [], [], []
    dia_index = {"nut": {}, "washer": {}}   # coating-aware component lookup for BOMs

    for sheet, cfg in SHEETS.items():
        if sheet not in wb.sheetnames:
            exceptions.append(f"Sheet missing: {sheet}"); continue
        rows = [list(r) for r in wb[sheet].iter_rows(values_only=True)]
        hi = find_header(rows)
        if hi is None:
            exceptions.append(f"{sheet}: header row not found, skipped"); continue
        header = rows[hi]
        kinds = {c: col_kind(h) for c, h in enumerate(header)}
        size_c = next((c for c,k in kinds.items() if k=="size"), None)
        code_c = next((c for c,k in kinds.items() if k=="code"), None)
        grade_c = next((c for c,k in kinds.items() if k=="grade"), None)
        ship_meta = {c: meta_lookup(rows, hi, c) for c,k in kinds.items() if k in ("intransit","ordered")}

        for r in rows[hi+1:]:
            size = clean(r[size_c]) if size_c is not None and size_c < len(r) else ""
            code = clean(r[code_c]) if code_c is not None and code_c < len(r) else ""
            if not size and not code: continue
            grade = clean(r[grade_c]) if grade_c is not None and grade_c < len(r) else cfg.get("grade","")
            if cfg.get("code_is_name"):
                sku = f"{cfg.get('prefix','X')}-{slug(code)}-{slug(size)}".strip("-")
                name = cfg["name"].format(code=code or cfg["category"], size=size, grade=grade)
            elif code:
                sku, name = code, cfg["name"].format(size=size, grade=grade)
            else:
                sku = f"{cfg.get('prefix', slug(sheet))}-{slug(size)}-{slug(grade)}".strip("-")
                name = cfg["name"].format(size=size, grade=grade)
            if not sku: continue

            if sku in items:
                exceptions.append(f"{sheet}: SKU {sku} already exists (from {items[sku]['source']}); quantities merged")
            else:
                items[sku] = dict(sku=sku, name=name, item_type=cfg["item_type"],
                                  category=cfg["category"], size=size, grade=grade,
                                  coating=cfg.get("coating",""), uom=cfg.get("uom","each"),
                                  source=sheet)
                dia = size.split("X")[0].strip()
                if cfg["category"]=="Nut" and cfg.get("coating")=="HDG" and sheet=="A563M HDG": dia_index["nut"][dia]=sku
                if cfg["category"]=="Washer" and sheet=="F436M HDG": dia_index["washer"][dia]=sku

            main_q = saif_q = 0; sales_cols = 0
            for c, k in kinds.items():
                v = r[c] if c < len(r) else None
                if not is_num(v) or v == 0: continue
                if k == "main": main_q += v
                elif k == "saif": saif_q += v
                elif k == "sales": sales_cols += v
                elif k in ("intransit","ordered"):
                    if v < 0:
                        exceptions.append(f"{sheet}: negative {k} {v} for {sku}, skipped"); continue
                    m = ship_meta.get(c, {})
                    shipments.append(dict(sku=sku, qty=int(v), status="in_transit" if k=="intransit" else "ordered",
                                          ref=m.get("ref") or f"{slug(sheet)[:10]}-C{c}", eta=m.get("eta"),
                                          party=m.get("party",""), source=sheet))
            for wh, q in (("WH-MAIN", main_q), ("WH-SAIF", saif_q)):
                if q < 0:
                    exceptions.append(f"{sheet}: negative on-hand {q} for {sku} at {wh}, clamped to 0"); q = 0
                if q: stock.append(dict(sku=sku, warehouse=wh, on_hand=int(q)))
            if sales_cols:
                exceptions.append(f"{sheet}: {sku} has {int(sales_cols)} in SALES/PO columns (not migrated as stock)")

    # Auto-generate BOMs for BNW / BNWW assembly sets
    for it in list(items.values()):
        cfg = SHEETS.get(it["source"], {})
        if it["item_type"] != "assembly": continue
        m = re.match(r"^A(A325MT1GC(M\d+)X\d+F)W?$", it["sku"])
        if not m:
            exceptions.append(f"BOM: could not parse assembly SKU {it['sku']}"); continue
        bolt_sku, dia = "B"+m.group(1), m.group(2)
        nut_sku, washer_sku = dia_index["nut"].get(dia), dia_index["washer"].get(dia)
        if bolt_sku not in items:
            items[bolt_sku] = dict(sku=bolt_sku, name=f"A325M T-1 Heavy Hex Full Thread Bolt HDG {it['size']}",
                                   item_type="part", category="Bolt", size=it["size"], grade="A325M",
                                   coating="HDG", uom="each", source="auto (BOM component)")
        if not nut_sku or not washer_sku:
            exceptions.append(f"BOM: missing nut/washer SKU for {it['sku']} (dia {dia}), BOM incomplete"); continue
        boms.append(dict(parent=it["sku"], components=[
            dict(sku=bolt_sku, qty=1), dict(sku=nut_sku, qty=1), dict(sku=washer_sku, qty=cfg.get("washers",1))]))

    seed = dict(
        generated=datetime.date.today().isoformat(),
        source="INVENTORY STOCK LIST AS ON 1 JULY 2026.xlsx",
        warehouses=[dict(id="WH-MAIN", name="Main Warehouse (Jebel Ali)", address="Jebel Ali, Dubai, UAE"),
                    dict(id="WH-SAIF", name="SAIF Zone Warehouse", address="SAIF Zone, Sharjah, UAE")],
        items=[{k:v for k,v in it.items()} for it in items.values()],
        stock=stock, shipments=shipments, boms=boms, exceptions=exceptions)
    with open(out_path, "w") as f: json.dump(seed, f, indent=1)
    print(f"items={len(items)} stock_rows={len(stock)} shipments={len(shipments)} boms={len(boms)} exceptions={len(exceptions)}")

if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2])
